# ------------------------------------------------------
# ----------------- datamanagement.py ------------------
# ------------------------------------------------------

### START IMPORTS ###
## Local imports ##
import src.sphcm  as sphcm
import src.sscm as sscm
import src.conversion as conversion
import src.functions as functions

from src.sphcm import *
from src.sscm import *
from src.conversion import *
from src.functions import *

## Openpyxl imports ##
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series, LineChart

## Matplotlib imports ##
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import (NavigationToolbar2QT as NavigationToolbar)
from matplotlib.backends.backend_qt5agg import FigureCanvas
from matplotlib.figure import Figure
from matplotlib.ticker import StrMethodFormatter

## Numpy imports ##
import numpy as np
from numpy import sum,isrealobj,sqrt
from numpy.random import standard_normal

## Lmfit imports and shapely ##
import lmfit
from lmfit import *
from shapely.geometry import LineString

## Other imports ##
import time
import random
import math
import sys
import json
import os
import csv
import re
import pprint
### END IMPORTS ###


### Other functions ###
## Auto-Process Functions ##
def autoProcessConvert3DEP(autoProcessUI):
    from_folder = autoProcessUI.fromFolderEntry.text()
    to_folder = autoProcessUI.toFolderEntry.text()

    files_list = os.listdir(from_folder)
    files_list_len = len(files_list)


    for file in files_list:
        i = os.path.join(from_folder, file)
        if os.path.isfile(i) and file.split('.')[1] == 'csv':
            if autoProcessUI.autoConvert3DEP.isChecked():
                load3DEP(i, to_folder)


def autoProcessLoadandFit(autoProcessUI, datapointsUI, ParamsUI, MainUI):
    try:
        ### UPDATE STATUS 1 ###
        autoProcessUI.statusLabel.setText('Computing...')
        autoProcessUI.autoProcessStartButton.setEnabled(False)
        autoProcessUI.closeAutoProcess.setEnabled(False)

        ### Run Auto Convert 3DEP ###
        autoProcessConvert3DEP(autoProcessUI)

        ### Main Function ###

        from_folder = autoProcessUI.fromFolderEntry.text()
        to_folder = autoProcessUI.toFolderEntry.text()

        if autoProcessUI.autoConvert3DEP.isChecked():
            files_list = os.listdir(to_folder)
        else:
            files_list = os.listdir(from_folder)

        files_list_len = len(files_list)

        what_to_fit = MainUI.what_to_fit
        method = MainUI.method
        trei_dep = MainUI.fitting_gen_fieldgrad_comboBox.currentText()

        local_params = MyParameters()
        if MainUI.modelMainTab.currentIndex() == 0:
            model = 'hopa'
            local_params.startupParameters(os.path.join(os.getcwd(), 'saves/default_hopa_parameters'))
        elif MainUI.modelMainTab.currentIndex() == 1:
            model = 'sish'
            local_params.startupParameters(os.path.join(os.getcwd(), 'saves/default_sish_parameters'))

        local_params.refreshFromGUIFittingParameters(ParamsUI, MainUI, model)

        data_objs_list = [Data() for i in range(len(files_list))]
        fittings_objs_list = [MyFittings() for i in range(len(files_list))]

        #### CENTRALIZES DATA 1 ####
        results_list = []
        fitted_values_dictionary = {'Fitted Frequency': [[],'',''],
                                'Fitted CMFactor': [[],'',''],
                                'Fitted DEPForce': [[],'','']}

        exp_values_dictionary = {'Experimental Frequency': [[], '', ''],
                                    'Experimental CMFactor': [[], '', ''],
                                    'Experimental DEPForce': [[], '', '']}

        validation_dictionary = {'Value status': [[], '', '']}

        CO_dictionary = {'1st CO frequency': [[],'','','', 'Hz', '1st CO frequency'],
                        '2nd CO frequency': [[],'','','', 'Hz', '2nd CO frequency']}

        etc_dict = {'Model': [[],model, 'None', '','','Model'],
                'Rsqare': [[], 'None', 'None', '','','Rsqare']}

        if model == 'hopa':
            local_dict = {'fitting_gen_buffer_cond': [[],'','','', 'S/m', 'Buffer Conductivity'],
                     'fitting_gen_buffer_perm': [[],'','','', 'ε', 'Buffer Permitivity'],
                     'fitting_hopa_particle_cond': [[],'','','', 'S/m', 'Particle Conductivity'],
                     'fitting_hopa_particle_perm': [[],'','','', 'ε', 'Particle Permitivity'],
                     'fitting_hopa_particle_radius': [[],'','','', 'μm', 'Particle Radius']}

        elif model == 'sish':
            local_dict = {'fitting_gen_buffer_cond': [[],'','','', 'S/m', 'Buffer Conductivity'],
                     'fitting_gen_buffer_perm': [[],'','','', 'ε', 'Buffer Permitivity'],
                     'fitting_sish_particle_radius': [[],'','','', 'μm', 'Particle Radius'],
                     'fitting_sish_membrane_thickness': [[],'','','', 'nm', 'Membrane Thickness'],
                     'fitting_sish_membrane_perm': [[],'','','', 'ε', 'Membrane Permitivity'],
                     'fitting_sish_membrane_cond': [[],'','','', 'S/m', 'Membrane Conductivity'],
                     'fitting_sish_cytoplasm_perm': [[],'','','', 'ε', 'Cytoplasm Permitivity'],
                     'fitting_sish_cytoplasm_cond': [[],'','','', 'S/m', 'Cytoplasm Conductivity']}

        ### FITS ALL DATA AND SAVES EACH IN EXCEL ###
        for file in files_list:
            index = files_list.index(file)
            if autoProcessUI.autoConvert3DEP.isChecked():
                i = os.path.join(to_folder, file)
            else:
                i = os.path.join(from_folder, file)


            if os.path.isfile(i) and file.split('.')[1] == 'xlsx':
                data_objs_list[index].loadExcelData(i)

                # OUTLAYERS MANAGEMENT
                if autoProcessUI.autoDetectOutliers.isChecked():
                    data_objs_list[index].dataValidation_list = fittings_objs_list[index].returnOutliers(frequency_list = data_objs_list[index].frequency_list,
                                                             CMfactor_list = data_objs_list[index].CMfactor_list,
                                                             DEPforce_list = data_objs_list[index].DEPforce_list,
                                                             threshold = float(autoProcessUI.outliersThresholdEntry.text())
                                                             )

                    #print(autoProcessUI.outliersThresholdEntry.text())

                    instanced_list2 =[]
                    for i in data_objs_list[index].dataValidation_list:
                        if i == 'Outlier':
                            instanced_list2.append('Disabled')
                        else:
                            instanced_list2.append('Enabled')
                    data_objs_list[index].dataStatus_list = instanced_list2

                # AUTO FIT DATA
                result = fittings_objs_list[index].fit(frequency_list = data_objs_list[index].frequency_list,
                                                       CMfactor_list = data_objs_list[index].CMfactor_list,
                                                       DEPforce_list = data_objs_list[index].DEPforce_list,
                                                       parameters = local_params.parameters,
                                                       model = model,
                                                       what_to_fit = what_to_fit,
                                                       trei_DEP = trei_dep,
                                                       method = method)

                data_objs_list[index].fittedFrequency_list, data_objs_list[index].fittedCMfactor_list, data_objs_list[index].fittedDEPforce_list = fittings_objs_list[index].returnFittedValues2(result, model, MainUI.what_to_fit, 50, data_objs_list[index].frequency_list, local_params)


                #CORRECTING DATA WITH FIELD GRADIENT
                if MainUI.fitting_gen_fieldgrad_comboBox.currentText() == 'Unspecified':
                    instanced_list = []
                    for i in data_objs_list[index].CMfactor_list:
                        instanced_list.append(
                            i / float(fittings_objs_list[index].returnFittedParamtersVar2(result, 'fitting_gen_fieldgrad')[0]))

                    instanced_errors_list = []
                    for j in data_objs_list[index].CMfactor_errors_list:
                        instanced_errors_list.append(
                            j / float(fittings_objs_list[index].returnFittedParamtersVar2(result, 'fitting_gen_fieldgrad')[0]))

                    data_objs_list[index].CMfactor_list = instanced_list
                    data_objs_list[index].CMfactor_errors_list = instanced_errors_list
                data_objs_list[index].saveExcelData(os.path.join(autoProcessUI.toFolderEntry.text(), file), model, MainUI, data_objs_list[index], fittings_objs_list[index])

                #### CENTRALIZES DATA 2 ####
                #print(fittings_objs_list[index].returnFitRsquare(result))

                if autoProcessUI.autoDetectRsquare.isChecked():
                    local_rsquare = float(fittings_objs_list[index].returnFitRsquare(result))
                else:
                    local_rsquare = 1.0

                if local_rsquare > float(autoProcessUI.rsquareThresholdEntry.text()):

                    results_list.append(result)
                    firstCO, secondCO, intersection = fittings_objs_list[index].cross_over_first(data_objs_list[index].fittedFrequency_list, data_objs_list[index].fittedCMfactor_list)
                    CO_dictionary['1st CO frequency'][0].append(firstCO)
                    CO_dictionary['2nd CO frequency'][0].append(secondCO)

                    fitted_values_dictionary['Fitted Frequency'][0].append(data_objs_list[index].fittedFrequency_list)
                    fitted_values_dictionary['Fitted CMFactor'][0].append(data_objs_list[index].fittedCMfactor_list)
                    fitted_values_dictionary['Fitted DEPForce'][0].append(data_objs_list[index].fittedDEPforce_list)

                    exp_values_dictionary['Experimental Frequency'][0].append(data_objs_list[index].frequency_list)
                    exp_values_dictionary['Experimental CMFactor'][0].append(data_objs_list[index].CMfactor_list)
                    exp_values_dictionary['Experimental DEPForce'][0].append(data_objs_list[index].DEPforce_list)
                    validation_dictionary['Value status'][0].append(data_objs_list[index].dataStatus_list)


        #### CENTRALIZES DATA 3 ####
            # Parameters #
        for index in range(len(results_list)):
            for i in local_dict.keys():
                local_dict[i][0].append(fittings_objs_list[index].returnFittedParamtersVar2(results_list[index], i)[0])

            # Cross-over #
        for i in CO_dictionary.keys():
            CO_dictionary[i][0] = [value for value in CO_dictionary[i][0] if value != 'N/A']

        CO_dictionary.update(local_dict)
        local_dict = CO_dictionary

            # Fitted data_01 #
        for i in fitted_values_dictionary.keys():
            fitted_list = np.stack(fitted_values_dictionary[i][0], axis=-1)
            average_list = []
            stdev_list = []
            for j in fitted_list:
                average = np.average(j)
                stdev = np.std(j)
                average_list.append(average)
                stdev_list.append(stdev)

            fitted_values_dictionary[i][1] = average_list
            fitted_values_dictionary[i][2] = stdev_list
            #print(i, ':  AVG:', fitted_values_dictionary[i][1], ' STD:', fitted_values_dictionary[i][2])

        # Experimental data_01 #
        indecies_list = []
        for i in validation_dictionary.keys():
            val_list = np.stack(validation_dictionary[i][0], axis=-1)
            x = 0
            for j in val_list:
                z = 0
                for y in j:
                    if y == "Disabled":
                        indecies_list.append([z, x])
                    z = z + 1
                x = x + 1

        for i in exp_values_dictionary.keys():
            exp_list = np.stack(exp_values_dictionary[i][0], axis=-1).tolist()

            try:
                for j in indecies_list:
                    column = int(j[0])
                    row = int(j[1])
                    exp_list[row][column] = 'N/A'
            except:
                print(' ')

            for j in range(len(exp_list)):
                exp_list[j] = [value for value in exp_list[j] if value != 'N/A']

            average_list = []
            stdev_list = []
            for j in exp_list:
                average = np.average(j)
                stdev = np.std(j)
                average_list.append(average)
                stdev_list.append(stdev)

            exp_values_dictionary[i][1] = average_list
            exp_values_dictionary[i][2] = stdev_list
            #print(i, ':  AVG:', exp_values_dictionary[i][1], ' STD:', exp_values_dictionary[i][2])

            # Parameters #
        for i in local_dict.keys():
            for j in range(0, len(local_dict[i][0])):
                local_dict[i][0][j] = float(local_dict[i][0][j])

            average = np.average(local_dict[i][0])
            stdev = np.std(local_dict[i][0])
            nosamples = len(local_dict[i][0])

            local_dict[i][1] = average
            local_dict[i][2] = stdev
            local_dict[i][3] = nosamples
            #print(i, ':  AVG:', local_dict[i][1], ' STD:', local_dict[i][2])

            # Adding etc dictionary #
        etc_dict.update(local_dict)
        local_dict = etc_dict
        #for i in local_dict.keys():
            #print(i, ':  AVG:', local_dict[i][1], ' STD:', local_dict[i][2])

            ### AUTO CREATE CENTRALIZED FILE ###
        if autoProcessUI.autoCreateCentralized.isChecked():
            centrfile = os.path.join(autoProcessUI.toFolderEntry.text(), 'Centralizer.xlsx')
            autoProcessMakeCentralizer(local_dict, exp_values_dictionary, fitted_values_dictionary, centrfile)

        ### UPDATE STATUS 2 ###
        autoProcessUI.statusLabel.setText('Finished!')
        autoProcessUI.autoProcessStartButton.setEnabled(True)
        autoProcessUI.closeAutoProcess.setEnabled(True)

    ### MANAGES ERRORS AND RESETS UI ###
    except:
        print("[ERROR] [Main] [AutoProcessUI] AutoProcess method could not run")
        autoProcessUI.statusLabel.setText('Failed!')
        autoProcessUI.autoProcessStartButton.setEnabled(True)
        autoProcessUI.closeAutoProcess.setEnabled(True)


def autoProcessMakeCentralizer(params_dict, exp_values_dict, fit_values_dict, file):
        instanced_data = Data()
        wb = Workbook()
        ws = wb.active

    ## Excel top-table setup ##
        ws['A1'] = 'Frequency (Hz)'
        ws['B1'] = 'Experimental CM factor'
        ws['C1'] = 'Experimental CM Factor errors'
        ws['D1'] = 'Experimental DEP force'
        ws['E1'] = 'Experimental DEP force errors'
        ws['F1'] = 'Status'
        ws['G1'] = 'Type'
        ws['H1'] = 'Fitted Frequency (Hz)'
        ws['I1'] = 'Fitted CMfactor'
        ws['J1'] = 'Fitted DEPforce'
        ws['K1'] = 'Name               '
        ws['L1'] = 'Value / Type'
        ws['M1'] = 'Error (+/-)'
        ws['N1'] = 'No.     '
        ws['O1'] = 'Unit    '

        instanced_data.autoStretchColumns(ws)
        instanced_data.setBackgroundColor(ws, top=1, bottom=1000, left=1, right=100, color="00EBF1DE")
        instanced_data.setBackgroundColor(ws, top=1, bottom=1, left=1, right=15, color='00C4D79B')
        instanced_data.setBorder(ws, top=1, bottom=1, left=1, right=15, color='000000', border_style='thick')

        x = 2
        for i in params_dict.keys():
            ws.cell(row=x, column=11, value=params_dict[i][5])
            if params_dict[i][5] == 'Model' or params_dict[i][5] == 'Rsqare':
                ws.cell(row=x, column=12, value=params_dict[i][1])
                ws.cell(row=x, column=13, value=params_dict[i][2])
                ws.cell(row=x, column=14, value=params_dict[i][3])
            else:
                ws.cell(row=x, column=12, value=float(params_dict[i][1]))
                ws.cell(row=x, column=13, value=float(params_dict[i][2]))
                ws.cell(row=x, column=14, value=float(params_dict[i][3]))
            ws.cell(row=x, column=15, value=params_dict[i][4])
            x = x + 1

        instanced_data.setBorder(ws, top=2, bottom=x-1, left=11, right=15, color='000000', border_style='double')
        instanced_data.setTabelBackground(ws, top=2, bottom=x-1, left=11, right=15, color1='00C0C0C0', color2='00FFFFFF')


        list = [exp_values_dict['Experimental Frequency'][1], exp_values_dict['Experimental CMFactor'][1], exp_values_dict['Experimental CMFactor'][2],exp_values_dict['Experimental DEPForce'][1], exp_values_dict['Experimental DEPForce'][2]]
        x = 1
        for i in list:
            y = 2
            for j in i:
                ws.cell(row=y, column=x, value=j)
                y = y + 1
            x = x + 1

        list2 = []
        for i in list:
            list2.append(len(i))
        instanced_data.setBorder(ws, top=2, bottom=max(list2)+1, left=1, right=7, color='000000', border_style='double')
        instanced_data.setTabelBackground(ws, top=2, bottom=max(list2)+1, left=1, right=7, color1='00C0C0C0', color2='00FFFFFF')


        list = fit_values_dict['Fitted Frequency'][1], fit_values_dict['Fitted CMFactor'][1], fit_values_dict['Fitted DEPForce'][1]
        x = 8
        for i in list:
            y = 2
            for j in i:
                ws.cell(row=y, column=x, value=j)
                y = y + 1
            x = x + 1

        list2 = []
        for i in list:
            list2.append(len(i))
        instanced_data.setBorder(ws, top=2, bottom=max(list2)+1, left=8, right=x - 1, color='000000', border_style='double')
        instanced_data.setTabelBackground(ws, top=2, bottom=max(list2)+1, left=8, right=x-1, color1='00C0C0C0', color2='00FFFFFF')

    # Adding the charts#
        instanced_data.chartPlotExcel(ws, what_to_plot="CMfactor", exp_freq=exp_values_dict['Experimental Frequency'][1],fit_freq=fit_values_dict['Fitted Frequency'][1], name='CM factor scatter', yname='CM factor', position="P1")
        instanced_data.chartPlotExcel(ws, what_to_plot="DEPforce", exp_freq=exp_values_dict['Experimental Frequency'][1],fit_freq=fit_values_dict['Fitted Frequency'][1], name='DEP force scatter', yname='DEP force', position="P20")

    # Saving and closing excel
        wb.save(filename=file)


## 3DEP management functions ##
def load3DEPloop(folder):
    for file in os.listdir(folder):
        f = os.path.join(folder, file)
        if os.path.isfile(f) and file.split('.')[1] == 'csv':
            print(f)
            load3DEP(f,None)

def load3DEP(fromFile, toFolder):
    wb = openpyxl.Workbook()
    ws = wb.active

    with open(fromFile) as f:
        reader = csv.reader(f, delimiter=':')
        for row in reader:
            ws.append(row)

    x = 1
    measured_frequency_list = []
    measured_DEPforce_list = []
    cells_deleted = []

    fitted_frequency_list = []
    fitted_DEPforce_list = []
    fitted_scaledDEPforce_list = []

    params = []

    while x < 100:
        if ws.cell(row=x, column=1).value == 'DEP-Spectrum':
            measuredSpectrum_row = x + 2
            print('Row for DEP-Spectrum is at row: ' + str(x))
        elif ws.cell(row=x, column=1).value == 'Auto-Fit Singel-Shell Modelparameters':
            print('Row for Model Spectrum is at row: ' + str(x + 15))
            modelSpectrum_row = x + 15
            modelParameters_row = x + 5
        elif ws.cell(row=x, column=1).value == 'Manual-Fit Singel-Shell Modelparameters':
            print('Row for Model Spectrum is at row: ' + str(x + 15))
            modelSpectrum_row = x + 15
            modelParameters_row = x + 5

        x = x + 1
    try:
        for i in range(modelParameters_row, modelParameters_row + 7):
            y = ws.cell(row=i, column=2).value
            chunks = re.split(', |S/|\(', y)
            y = ws.cell(row=i, column=1).value

            params.append([y, float(chunks[1].replace(',', '.'))])

        print(params)
    except:
        print('[WARNING] [datamanagement] [load3DEP] Data was never fitted / No parameters to convert')

    for i in range(measuredSpectrum_row, measuredSpectrum_row + 20):
        y = ws.cell(row=i, column=1).value
        chunks = y.split(' ,')
        print(chunks[0].replace(',', '.'), chunks[1].replace(',', '.'))
        measured_frequency_list.append(chunks[0].replace(',', '.'))
        measured_DEPforce_list.append(chunks[1].replace(',', '.'))

        try:
            if chunks[2] == ' Point removed from analysis / Manually removed  Outlier  ':
                cells_deleted.append(chunks[2])
            elif chunks[2] == ' Point removed from analysis / Manually removed  ':
                cells_deleted.append(chunks[2])


        except:
            print('test')
    #print(modelSpectrum_row, modelSpectrum_row + 20 - len(cells_deleted))
    #print(len(cells_deleted))
    try:
        for i in range(modelSpectrum_row, modelSpectrum_row + 20 - len(cells_deleted)):
            y = ws.cell(row=i, column=1).value

            chunks = y.split(' ,')
            print(chunks[0].replace(',','.'), chunks[1].replace(',','.'), chunks[2].replace(',','.'))
            fitted_frequency_list.append(chunks[0].replace(',','.'))
            fitted_DEPforce_list.append(chunks[1].replace(',','.'))
            fitted_scaledDEPforce_list.append(chunks[2].replace(',','.'))
    except:
        print('[WARNING] [datamanagement] [load3DEP] Data was never fitted / No fitted values to convert')


    wb = openpyxl.Workbook()
    ws = wb.active
    instanced_data = Data()

    ws['A1'] = 'Frequency (Hz)'
    ws['B1'] = 'Experimental CM factor'
    ws['C1'] = 'Experimental CM Factor errors'
    ws['D1'] = 'Experimental DEP force'
    ws['E1'] = 'Experimental DEP force errors'
    ws['F1'] = 'Status'
    ws['G1'] = 'Type'
    ws['H1'] = 'Fitted Frequency (Hz)'
    ws['I1'] = 'Fitted CMfactor'
    ws['J1'] = 'Fitted DEPforce'
    ws['K1'] = 'Name               '
    ws['L1'] = 'Value / Type'
    ws['M1'] = 'Error (+/-)'
    ws['N1'] = 'Unit     '

    instanced_data.autoStretchColumns(ws)
    instanced_data.setBackgroundColor(ws, top=1, bottom=1000, left=1, right=100, color="00EBF1DE")
    instanced_data.setBackgroundColor(ws, top=1, bottom=1, left=1, right=14, color='00C4D79B')
    instanced_data.setBorder(ws, top=1, bottom=1, left=1, right=14, color='000000', border_style='thick')

    list = [measured_frequency_list, measured_DEPforce_list]
    x = 1
    for i in list:
        y = 2
        for j in i:
            ws.cell(row=y, column=x, value=float(j))
            y = y + 1
        x = x + 1

    list2 = []
    for i in list:
        list2.append(len(i))
    instanced_data.setBorder(ws, top=2, bottom=max(list2)+1, left=1, right=7, color='000000', border_style='double')
    instanced_data.setTabelBackground(ws, top=2, bottom=max(list2)+1, left=1, right=7, color1='00C0C0C0', color2='00FFFFFF')

    try:
        list = [fitted_frequency_list, fitted_DEPforce_list]
        x = 8
        for i in list:
            y = 2
            for j in i:
                ws.cell(row=y, column=x, value=float(j))
                y = y + 1
            x = x + 1

        list2 = []
        for i in list:
            list2.append(len(i))
        instanced_data.setBorder(ws, top=2, bottom=max(list2) + 1, left=8, right=x, color='000000', border_style='double')
        instanced_data.setTabelBackground(ws, top=2, bottom=max(list2) + 1, left=8, right=x, color1='00C0C0C0', color2='00FFFFFF')

        ws.cell(row=2, column=14).value = 'S/m'
        ws.cell(row=3, column=14).value = 'ε'
        ws.cell(row=4, column=14).value = 'S/m^2'
        ws.cell(row=5, column=14).value = 'F^/m^2'
        ws.cell(row=6, column=14).value = 'μm'
        ws.cell(row=7, column=14).value = 'S/m'
        ws.cell(row=8, column=14).value = 'ε'

        for i in params:
            ws.cell(row=params.index(i) + 2, column=11).value = i[0]
            ws.cell(row=params.index(i) + 2, column=12).value = i[1]

        instanced_data.setBorder(ws, top=2, bottom=8, left=11, right=14  , color='000000', border_style='double')
        instanced_data.setTabelBackground(ws, top=2, bottom=8, left=11, right=14, color1='00C0C0C0', color2='00FFFFFF')
    except:
        print('\n')

    file = fromFile.replace(".csv", ".xlsx", 1)
    if toFolder != None:
        file = os.path.join(toFolder, file.split('\\')[-1])
        print(file)
    wb.save(file)

    ### Excel module ###
def addExcel(wb, colnumb,frequencyList, CMfactorList, CMfactorNoiseList, DEPforceList, DEPforceNoiseList, paramsList, model, with_noise):
    sheet = wb.active

    if model == 'hopa':
        sheet.cell(row=1, column=colnumb + 1).value = "Electric field gradient (V/mm2)"
        sheet.cell(row=2, column=colnumb + 1).value = "Particle radius (μm)"
        sheet.cell(row=3, column=colnumb + 1).value = "Particle permitivity (ε)"
        sheet.cell(row=4, column=colnumb + 1).value = "Particle conductivity (S/m)"
        sheet.cell(row=5, column=colnumb + 1).value = "Buffer permitivity (ε)"
        sheet.cell(row=6, column=colnumb + 1).value = "Buffer conductivity (S/m)"
        z = 8

    elif model == 'sish':
        sheet.cell(row=1, column=colnumb + 1).value = "Electric field gradient (V/mm2)"
        sheet.cell(row=2, column=colnumb + 1).value = "Particle radius (μm)"
        sheet.cell(row=3, column=colnumb + 1).value = "Membrane thickness (nm)"
        sheet.cell(row=4, column=colnumb + 1).value = "Membrane permitivity (ε)"
        sheet.cell(row=5, column=colnumb + 1).value = "Membrane conductivity (S/m)"
        sheet.cell(row=6, column=colnumb + 1).value = "Cytoplasm permitivity (ε)"
        sheet.cell(row=7, column=colnumb + 1).value = "Cytoplasm conductivity (S/m)"
        sheet.cell(row=8, column=colnumb + 1).value = "Buffer permitivity (ε)"
        sheet.cell(row=9, column=colnumb + 1).value = "Buffer conductivity (S/m)"
        z = 11

    y = 1
    for j in paramsList:
        sheet.cell(row=y, column=colnumb + 2).value = j
        y = y + 1

    y = z
    sheet.cell(row=y, column=colnumb).value = "Frequency (Hz)"
    for i in frequencyList:
        sheet.cell(row=y+1, column=colnumb).value = i
        y = y + 1

    y = z
    sheet.cell(row=y, column=colnumb + 1).value = "CM factor"
    for j in CMfactorList:
        sheet.cell(row=y+1, column=colnumb + 1).value = j
        y = y + 1

    y = z
    sheet.cell(row=y, column=colnumb + 2).value = "DEP force"
    for j in DEPforceList:
        sheet.cell(row=y+1, column=colnumb + 2).value = j
        y = y + 1

    if with_noise == True:
        y = z
        sheet.cell(row=y, column=colnumb + 3).value = "CM factor with noise"
        for j in CMfactorNoiseList:
            sheet.cell(row=y + 1, column=colnumb + 3).value = j
            y = y + 1

        y = z
        sheet.cell(row=y, column=colnumb + 4).value = "DEP force with noise"
        for j in DEPforceNoiseList:
            sheet.cell(row=y + 1, column=colnumb + 4).value = j
            y = y + 1

def frequencyConvertor(value, division):
    if division.currentText() == 'Hz':
        x = int(value.text())

    elif division.currentText() == 'kHz':
        x = int(value.text()) * 1000

    elif division.currentText() == 'MHz':
        x = int(value.text()) * 1000000

    return x


def awgnNoise(s, SNRdB, L=1):
#AWGN channel
#Add AWGN noise to output signal. The function adds AWGN noise vector to signal 's' to generate a resulting signal vector 'r' of specified SNR in dB. It also
#returns the noise vector 'n' that is added to the signal 's' and the power spectral density N0 of noise added
#Parameters:
    #s : output/transmitted signal vector
    #SNRdB : desired signal to noise ratio (expressed in dB) for the received signal
    #L : oversampling factor (applicable for waveform simulation) default L = 1.
#Returns:
    #r : received signal vector (r=s+n)
    s = np.array(s)
    gamma = 10 ** (SNRdB / 10)  # SNR to linear scale
    if s.ndim == 1:  # if s is single dimensional vector
        P = L * sum(abs(s) ** 2) / len(s)  # Actual power in the vector
    else:  # multi-dimensional signals like MFSK
        P = L * sum(sum(abs(s) ** 2)) / len(s)  # if s is a matrix [MxN]
    N0 = P / gamma  # Find the noise spectral density
    if isrealobj(s):  # check if output is real/complex object type
        n = sqrt(N0 / 2) * standard_normal(s.shape)  # computed noise
    else:
        n = sqrt(N0 / 2) * (standard_normal(s.shape) + 1j * standard_normal(s.shape))
    r = s + n  # received signal
    return r


    ### Fitting module ###
class MyParameters():

    def __init__(self):
        # Instance Variable #
        self.parameters = lmfit.Parameters()

    ## Parameter methods ###
# Loads last used parameters #
    def startupParameters(self, file):
        self.loadfile = open(file, "r")
        self.parameters.load(fp=self.loadfile)
        self.loadfile.close()

# Loads parameters from file #
    def loadParameters(self, file):
        self.loadfile = open(file, "r")
        self.parameters.load(fp=self.loadfile)
        self.loadfile.close()

# Saves parameters to file #
    def saveParameters(self, file):
        self.savefile = open(file, "w")
        print(self.savefile)
        self.parameters.dump(fp=self.savefile)
        self.savefile.close()

# Returns all parameters in a dict #
    def returnParameters(self):
        return self.parameters

# Returns a specific info about a parameter #
    def returnParameter(self, parameter, info):
        parameterString = str(self.parameters.get(parameter))

        if info == 'value' or 'vary':
            start = parameterString.index('value') + len('value') + 1
            end = parameterString.index(',', parameterString.index('value') + len('value') + 1)
            x = parameterString[start:end]

            if info == 'value':
                if '(fixed)' in x:
                    return x[0:x.index('(fixed)')]
                else:
                    return x

            elif info == 'vary':
                if '(fixed)' in x:
                    return False
                else:
                    return True

        if info == 'min' or 'max':
            start = parameterString.index('bounds') + len('bounds') + 1
            x = parameterString[start:len(parameterString)-1]

            if info == 'min':
                return x[x.index('[')+1:x.index(':')]

            elif info == 'max':
                return x[x.index(':')+1:x.index(']')]

#Returns a dictionary with all info regarding a parameter
    def returnParameterDict(self, parameter):
        parameterString = str(self.parameters.get(parameter))
        start = parameterString.index('value') + len('value') + 1
        end = parameterString.index(',', parameterString.index('value') + len('value') + 1)
        x = parameterString[start:end]

        if '(fixed)' in x:
            valueParam = x[0:x.index('(fixed)')]
        else:
            valueParam = x

        if '(fixed)' in x:
            varyParam = False
        else:
            varyParam = True

        start = parameterString.index('bounds') + len('bounds') + 1
        x = parameterString[start:len(parameterString)-1]
        minParam = x[x.index('[')+1:x.index(':')]
        maxParam = x[x.index(':')+1:x.index(']')]

        parameterDict = {'Value':valueParam, 'Vary':varyParam, 'Min':minParam, 'Max':maxParam}
        return parameterDict

#Changes a specific info of a parameter
    def changeParameter(self, parameter, info, value):
        if info == 'value':
            x = self.parameters.get(parameter)
            x.set(value=value)
        elif info == 'vary':
            x = self.parameters.get(parameter)
            x.set(vary=value)
        elif info == 'min':
            x = self.parameters.get(parameter)
            x.set(min=value)
        elif info == 'max':
            x = self.parameters.get(parameter)
            x.set(max=value)

        return x

#Changes all values of a parameter
    def changeParameterDict(self, parameter, **kwargs):
        x = self.parameters.get(parameter)
        for y in kwargs.keys():
            if y == 'value':
                x.set(value=kwargs['value'])
            elif y == 'vary':
                x.set(vary=kwargs['vary'])
            elif y == 'min':
                x.set(value=kwargs['min'])
            elif y == 'max':
                x.set(vary=kwargs['max'])


## Display parameter settings in GUI ##
# Updating parameters in GUI #
    def changeGUIFittingParameters(self, ParametersUI, MainUI, model):

        if model == 'hopa':
            listoflists = [
                            ['fitting_gen_fieldgrad', [MainUI.fitting_gen_fieldgrad_value_entry]],
                            ['fitting_gen_buffer_cond', [MainUI.fitting_gen_buffer_cond_value_entry, MainUI.fitting_gen_buffer_cond_vary, ParametersUI.fitting_gen_buffer_cond_min_entry, ParametersUI.fitting_gen_buffer_cond_max_entry]],
                            ['fitting_gen_buffer_perm', [MainUI.fitting_gen_buffer_perm_value_entry, MainUI.fitting_gen_buffer_perm_vary, ParametersUI.fitting_gen_buffer_perm_min_entry, ParametersUI.fitting_gen_buffer_perm_max_entry]],

                            ['fitting_hopa_particle_cond', [MainUI.fitting_hopa_particle_cond_value_entry, MainUI.fitting_hopa_particle_cond_vary, ParametersUI.fitting_hopa_particle_cond_min_entry, ParametersUI.fitting_hopa_particle_cond_max_entry]],
                            ['fitting_hopa_particle_perm', [MainUI.fitting_hopa_particle_perm_value_entry, MainUI.fitting_hopa_particle_perm_vary, ParametersUI.fitting_hopa_particle_perm_min_entry, ParametersUI.fitting_hopa_particle_perm_max_entry]],
                            ['fitting_hopa_particle_radius', [MainUI.fitting_hopa_particle_radius_value_entry, MainUI.fitting_hopa_particle_radius_vary, ParametersUI.fitting_hopa_particle_radius_min_entry, ParametersUI.fitting_hopa_particle_radius_max_entry]],
                            ]

        elif model == 'sish':
            listoflists = [
                            ['fitting_gen_fieldgrad', [MainUI.fitting_gen_fieldgrad_value_entry]],
                            ['fitting_gen_buffer_cond', [MainUI.fitting_gen_buffer_cond_value_entry, MainUI.fitting_gen_buffer_cond_vary, ParametersUI.fitting_gen_buffer_cond_min_entry, ParametersUI.fitting_gen_buffer_cond_max_entry]],
                            ['fitting_gen_buffer_perm', [MainUI.fitting_gen_buffer_perm_value_entry, MainUI.fitting_gen_buffer_perm_vary, ParametersUI.fitting_gen_buffer_perm_min_entry, ParametersUI.fitting_gen_buffer_perm_max_entry]],

                            ['fitting_sish_particle_radius', [MainUI.fitting_sish_particle_radius_value_entry, MainUI.fitting_sish_particle_radius_vary, ParametersUI.fitting_sish_particle_radius_min_entry, ParametersUI.fitting_sish_particle_radius_max_entry]],
                            ['fitting_sish_membrane_thickness', [MainUI.fitting_sish_membrane_thickness_value_entry, MainUI.fitting_sish_membrane_thickness_vary, ParametersUI.fitting_sish_membrane_thickness_min_entry, ParametersUI.fitting_sish_membrane_thickness_max_entry]],

                            ['fitting_sish_membrane_perm', [MainUI.fitting_sish_membrane_perm_value_entry, MainUI.fitting_sish_membrane_perm_vary, ParametersUI.fitting_sish_membrane_perm_min_entry, ParametersUI.fitting_sish_membrane_perm_max_entry]],
                            ['fitting_sish_membrane_cond', [MainUI.fitting_sish_membrane_cond_value_entry, MainUI.fitting_sish_membrane_cond_vary, ParametersUI.fitting_sish_membrane_cond_min_entry, ParametersUI.fitting_sish_membrane_cond_max_entry]],

                            ['fitting_sish_cytoplasm_perm', [MainUI.fitting_sish_cytoplasm_perm_value_entry, MainUI.fitting_sish_cytoplasm_perm_vary, ParametersUI.fitting_sish_cytoplasm_perm_min_entry, ParametersUI.fitting_sish_cytoplasm_perm_max_entry]],
                            ['fitting_sish_cytoplasm_cond', [MainUI.fitting_sish_cytoplasm_cond_value_entry, MainUI.fitting_sish_cytoplasm_cond_vary, ParametersUI.fitting_sish_cytoplasm_cond_min_entry, ParametersUI.fitting_sish_cytoplasm_cond_max_entry]],
                            ]

        for x in listoflists:
            x[1][0].setText(str(self.returnParameter(x[0], 'value')))
            try:
                if self.returnParameter(x[0], 'vary') == True:
                    x[1][1].setCurrentText('Vary')
                else:
                    x[1][1].setCurrentText('Fixed')

                x[1][2].setText(str(self.returnParameter(x[0], 'min')))
                x[1][3].setText(str(self.returnParameter(x[0], 'max')))
            except:
                print(' ')

# Updating parameters from GUI #
    def refreshFromGUIFittingParameters(self, ParametersUI, MainUI, model):

        if model == 'hopa':
            listoflists = [
                            ['fitting_gen_fieldgrad', [MainUI.fitting_gen_fieldgrad_value_entry]],
                            ['fitting_gen_buffer_cond', [MainUI.fitting_gen_buffer_cond_value_entry, MainUI.fitting_gen_buffer_cond_vary, ParametersUI.fitting_gen_buffer_cond_min_entry, ParametersUI.fitting_gen_buffer_cond_max_entry]],
                            ['fitting_gen_buffer_perm', [MainUI.fitting_gen_buffer_perm_value_entry, MainUI.fitting_gen_buffer_perm_vary, ParametersUI.fitting_gen_buffer_perm_min_entry, ParametersUI.fitting_gen_buffer_perm_max_entry]],

                            ['fitting_hopa_particle_cond', [MainUI.fitting_hopa_particle_cond_value_entry, MainUI.fitting_hopa_particle_cond_vary, ParametersUI.fitting_hopa_particle_cond_min_entry, ParametersUI.fitting_hopa_particle_cond_max_entry]],
                            ['fitting_hopa_particle_perm', [MainUI.fitting_hopa_particle_perm_value_entry, MainUI.fitting_hopa_particle_perm_vary, ParametersUI.fitting_hopa_particle_perm_min_entry, ParametersUI.fitting_hopa_particle_perm_max_entry]],
                            ['fitting_hopa_particle_radius', [MainUI.fitting_hopa_particle_radius_value_entry, MainUI.fitting_hopa_particle_radius_vary, ParametersUI.fitting_hopa_particle_radius_min_entry, ParametersUI.fitting_hopa_particle_radius_max_entry]],
                            ]

        elif model == 'sish':
            listoflists = [
                            ['fitting_gen_fieldgrad', [MainUI.fitting_gen_fieldgrad_value_entry]],
                            ['fitting_gen_buffer_cond', [MainUI.fitting_gen_buffer_cond_value_entry, MainUI.fitting_gen_buffer_cond_vary, ParametersUI.fitting_gen_buffer_cond_min_entry, ParametersUI.fitting_gen_buffer_cond_max_entry]],
                            ['fitting_gen_buffer_perm', [MainUI.fitting_gen_buffer_perm_value_entry, MainUI.fitting_gen_buffer_perm_vary, ParametersUI.fitting_gen_buffer_perm_min_entry, ParametersUI.fitting_gen_buffer_perm_max_entry]],

                            ['fitting_sish_particle_radius', [MainUI.fitting_sish_particle_radius_value_entry, MainUI.fitting_sish_particle_radius_vary, ParametersUI.fitting_sish_particle_radius_min_entry, ParametersUI.fitting_sish_particle_radius_max_entry]],
                            ['fitting_sish_membrane_thickness', [MainUI.fitting_sish_membrane_thickness_value_entry, MainUI.fitting_sish_membrane_thickness_vary, ParametersUI.fitting_sish_membrane_thickness_min_entry, ParametersUI.fitting_sish_membrane_thickness_max_entry]],

                            ['fitting_sish_membrane_perm', [MainUI.fitting_sish_membrane_perm_value_entry, MainUI.fitting_sish_membrane_perm_vary, ParametersUI.fitting_sish_membrane_perm_min_entry, ParametersUI.fitting_sish_membrane_perm_max_entry]],
                            ['fitting_sish_membrane_cond', [MainUI.fitting_sish_membrane_cond_value_entry, MainUI.fitting_sish_membrane_cond_vary, ParametersUI.fitting_sish_membrane_cond_min_entry, ParametersUI.fitting_sish_membrane_cond_max_entry]],

                            ['fitting_sish_cytoplasm_perm', [MainUI.fitting_sish_cytoplasm_perm_value_entry, MainUI.fitting_sish_cytoplasm_perm_vary, ParametersUI.fitting_sish_cytoplasm_perm_min_entry, ParametersUI.fitting_sish_cytoplasm_perm_max_entry]],
                            ['fitting_sish_cytoplasm_cond', [MainUI.fitting_sish_cytoplasm_cond_value_entry, MainUI.fitting_sish_cytoplasm_cond_vary, ParametersUI.fitting_sish_cytoplasm_cond_min_entry, ParametersUI.fitting_sish_cytoplasm_cond_max_entry]],
                            ]

        for x in listoflists:

            self.changeParameter(x[0], 'value', float(x[1][0].text()))
            try:
                if x[1][1].currentText() == 'Vary':
                    self.changeParameter(x[0], 'vary', True)
                else:
                    self.changeParameter(x[0], 'vary', False)

                if float(x[1][2].text()) < float(x[1][3].text()):
                    self.changeParameter(x[0], 'min', float(x[1][2].text()))
                    self.changeParameter(x[0], 'max', float(x[1][3].text()))
            except:
                print(' ')

    def changeClassicalToElectrical(self, fromWhat, toWhat, parameter, UI):
        if parameter == 'conductance':
            toWhat.setText(str(round(
                float(fromWhat.text()) / float(
                    UI.fitting_sish_membrane_thickness_value_entry.text()) * 1E9, 10)))

        elif parameter == 'capacitance':
            toWhat.setText(str(round(
                float(fromWhat.text()) / float(
                    UI.fitting_sish_membrane_thickness_value_entry.text()) * 8.84194, 10)))

    def changeElectricalToClassical(self, fromWhat, toWhat, parameter, UI):
        if parameter == 'conductivity':
            toWhat.setText(str(round(
                float(fromWhat.text()) * float(
                    UI.fitting_sish_membrane_thickness_value_entry.text()) / 1E9, 10)))

        elif parameter == 'permitivity':
            toWhat.setText(str(round(
                float(fromWhat.text()) * float(
                    UI.fitting_sish_membrane_thickness_value_entry.text()) / 8.84194, 10)))

## Fit methods ##
class MyFittings():

    def fit(self, frequency_list, CMfactor_list, DEPforce_list, parameters, model, what_to_fit, trei_DEP, method):
        if trei_DEP == 'Unspecified' and model == 'hopa' and what_to_fit == 'CMfactor':
            self.mod = Model(hopa_CMfactor_3DEP)
        elif trei_DEP == 'Unspecified' and model == 'sish' and what_to_fit == 'CMfactor':
            self.mod = Model(sish_CMfactor_3DEP)
        elif model == 'hopa' and what_to_fit == 'CMfactor':
            self.mod = Model(hopa_CMfactor)
        elif model == 'hopa' and what_to_fit == 'DEPforce':
            self.mod = Model(hopa_DEPforce)
        elif model ==  'sish'and what_to_fit == 'CMfactor':
            self.mod = Model(sish_CMfactor)
        elif model ==  'sish'and what_to_fit == 'DEPforce':
            self.mod = Model(sish_DEPforce)

        if what_to_fit == 'CMfactor':
            self.result = self.mod.fit(CMfactor_list, parameters, freq=frequency_list, method=method, max_nfev=5000)
        elif what_to_fit == 'DEPforce':
            self.result = self.mod.fit(DEPforce_list, parameters, freq=frequency_list, method=method, max_nfev=5000)

        return self.result

    def printFit(self, result):
        print(fit_report(result))


    def returnFitRsquare(self, result):
        self.best_fit = result.best_fit
        self.residual = result.residual
        return 1.0 - (sum(self.residual ** 2.0) / sum(self.best_fit ** 2.0))


    def returnChisquare(self, result):
        return result.chisqr

    def returnOutliers(self, frequency_list, CMfactor_list, DEPforce_list, threshold):

        print(os.getcwd())
        outliersParameters = MyParameters()
        outliersParameters.loadParameters(os.path.join(os.getcwd(), 'saves/default_outliers_parameters'))
        validation_list = []
        print("Hello2")
        if len(CMfactor_list) > 2:
            what_to_fit = 'CMfactor'
        elif len(DEPforce_list) > 2:
            what_to_fit = 'DEPforce'

        self.result = self.fit(frequency_list=frequency_list, CMfactor_list=CMfactor_list, DEPforce_list=DEPforce_list, parameters=outliersParameters.parameters, model='sish', what_to_fit=what_to_fit, trei_DEP='Unspecified', method='powell')
        residuals = self.result.residual
        #print(residuals)


        mean_1 = np.mean(residuals)
        std_1 = np.std(residuals)
        #print(mean_1)
        #print(std_1)

        validation_list.append("Normal")
        for i in range(1,len(residuals)-1):
            #print(i)
            z_score = (residuals[i] - mean_1) / std_1
            #print(z_score)
            if np.abs(z_score) > threshold:
                validation_list.append("Outlier")
            else:
                validation_list.append("Normal")
        validation_list.append("Normal")
        return validation_list


    def returnFittedParamtersVar2(self, result, parameter):
        try:
            string_dict = {}
            start = result.fit_report().index('[[Variables]]')
            try:
                stop = result.fit_report().index('[[Correlations]]')
                string = result.fit_report()[start + 14:stop]
            except:
                string = result.fit_report()[start + 14:]

            string_list = string.splitlines()
            for i in string_list:
                j = i.split()
                if len(j) == 3 or len(j) == 5:
                    string_dict[j[0].replace(':','')] = [j[1], 'None']
                elif len(j) == 8:
                    string_dict[j[0].replace(':','')] = [j[1], j[3]]

            return string_dict.get(parameter)

        except:
            print("[ERROR] [datamanagement] [MyFittings] returnFittedParamtersVar2 method could not run")


    def returnFittedParameters(self, result, MainUI, model):
        dict1 = result.best_values

        if model == 'hopa':
            dict2 = {'fitting_gen_buffer_cond': [MainUI.fitting_gen_buffer_cond_value_entry, MainUI.fitting_gen_buffer_cond_vary],
                    'fitting_gen_buffer_perm': [MainUI.fitting_gen_buffer_perm_value_entry, MainUI.fitting_gen_buffer_perm_vary],
                    'fitting_hopa_particle_cond': [MainUI.fitting_hopa_particle_cond_value_entry, MainUI.fitting_hopa_particle_cond_vary],
                    'fitting_hopa_particle_perm': [MainUI.fitting_hopa_particle_perm_value_entry, MainUI.fitting_hopa_particle_perm_vary],
                    'fitting_hopa_particle_radius': [MainUI.fitting_hopa_particle_radius_value_entry, MainUI.fitting_hopa_particle_radius_vary]}

        elif model == 'sish':
            dict2 = {'fitting_gen_buffer_cond': [MainUI.fitting_gen_buffer_cond_value_entry, MainUI.fitting_gen_buffer_cond_vary],
                    'fitting_gen_buffer_perm': [MainUI.fitting_gen_buffer_perm_value_entry, MainUI.fitting_gen_buffer_perm_vary],
                    'fitting_sish_particle_radius': [MainUI.fitting_sish_particle_radius_value_entry, MainUI.fitting_sish_particle_radius_vary],
                    'fitting_sish_membrane_thickness': [MainUI.fitting_sish_membrane_thickness_value_entry, MainUI.fitting_sish_membrane_thickness_vary],

                    'fitting_sish_membrane_perm': [MainUI.fitting_sish_membrane_perm_value_entry, MainUI.fitting_sish_membrane_perm_vary],
                    'fitting_sish_membrane_cond': [MainUI.fitting_sish_membrane_cond_value_entry, MainUI.fitting_sish_membrane_cond_vary],

                    'fitting_sish_cytoplasm_perm': [MainUI.fitting_sish_cytoplasm_perm_value_entry, MainUI.fitting_sish_cytoplasm_perm_vary],
                    'fitting_sish_cytoplasm_cond': [MainUI.fitting_sish_cytoplasm_cond_value_entry, MainUI.fitting_sish_cytoplasm_cond_vary]}


        for key in dict1:
            try:
                if dict2[key][1].currentText() == 'Vary':
                    dict2[key][0].setText(str(round(dict1[key],10)))
            except:
                print(' ')

    def returnFittedValues(self, result):
        return result.best_fit

    def returnFittedValues2(self, result, model, what_to_fit, points, frequency, params):
        fittedFrequency_list = np.logspace(np.log10(frequency[0] - 0.3 * frequency[0]), np.log10(frequency[-1] + 0.5 * frequency[-1]), num=points, base=10)
        fittedCMfactor_list = []
        fittedDEPforce_list = []
        dict1 = result.best_values

        for i in fittedFrequency_list:
            if model == 'hopa' and what_to_fit == 'CMfactor':
                j = hopa_CMfactor(i, dict1['fitting_hopa_particle_perm'], dict1['fitting_hopa_particle_cond'], dict1['fitting_gen_buffer_perm'], dict1['fitting_gen_buffer_cond'])
                y = hopa_DEPforce(i, float(params.returnParameter('fitting_gen_fieldgrad', 'value')), float(params.returnParameter('fitting_hopa_particle_radius', 'value')), dict1['fitting_hopa_particle_perm'], dict1['fitting_hopa_particle_cond'], dict1['fitting_gen_buffer_perm'], dict1['fitting_gen_buffer_cond'])

            elif model == 'hopa' and what_to_fit == 'DEPforce':
                j = hopa_CMfactor(i, dict1['fitting_hopa_particle_perm'], dict1['fitting_hopa_particle_cond'], dict1['fitting_gen_buffer_perm'], dict1['fitting_gen_buffer_cond'])
                y = hopa_DEPforce(i, float(params.returnParameter('fitting_gen_fieldgrad', 'value')), dict1['fitting_hopa_particle_radius'], dict1['fitting_hopa_particle_perm'], dict1['fitting_hopa_particle_cond'], dict1['fitting_gen_buffer_perm'], dict1['fitting_gen_buffer_cond'])

            elif model == 'sish' and what_to_fit == 'CMfactor':
                j = sish_CMfactor(i, dict1['fitting_sish_particle_radius'], dict1['fitting_sish_membrane_thickness'], dict1['fitting_sish_membrane_perm'], dict1['fitting_sish_membrane_cond'], dict1['fitting_sish_cytoplasm_perm'], dict1['fitting_sish_cytoplasm_cond'], dict1['fitting_gen_buffer_perm'], dict1['fitting_gen_buffer_cond'])
                y = sish_DEPforce(i, float(params.returnParameter('fitting_gen_fieldgrad', 'value')), dict1['fitting_sish_particle_radius'], dict1['fitting_sish_membrane_thickness'], dict1['fitting_sish_membrane_perm'], dict1['fitting_sish_membrane_cond'], dict1['fitting_sish_cytoplasm_perm'], dict1['fitting_sish_cytoplasm_cond'], dict1['fitting_gen_buffer_perm'], dict1['fitting_gen_buffer_cond'])

            elif model == 'sish' and what_to_fit == 'DEPforce':
                j = sish_CMfactor(i, dict1['fitting_sish_particle_radius'], dict1['fitting_sish_membrane_thickness'], dict1['fitting_sish_membrane_perm'], dict1['fitting_sish_membrane_cond'], dict1['fitting_sish_cytoplasm_perm'], dict1['fitting_sish_cytoplasm_cond'], dict1['fitting_gen_buffer_perm'], dict1['fitting_gen_buffer_cond'])
                y = sish_DEPforce(i, float(params.returnParameter('fitting_gen_fieldgrad', 'value')), dict1['fitting_sish_particle_radius'], dict1['fitting_sish_membrane_thickness'], dict1['fitting_sish_membrane_perm'], dict1['fitting_sish_membrane_cond'], dict1['fitting_sish_cytoplasm_perm'], dict1['fitting_sish_cytoplasm_cond'], dict1['fitting_gen_buffer_perm'], dict1['fitting_gen_buffer_cond'])


            fittedCMfactor_list.append(j)
            fittedDEPforce_list.append(y)

        return fittedFrequency_list, fittedCMfactor_list, fittedDEPforce_list

    def returnFittedResiduals(self, result):
        return result.residual


# Calculating the cross_over #
    def cross_over_first(self, frequencyList, CMfactorList):
        zeroList = [0] * len(frequencyList)
        backList = [0.01] * len(frequencyList)

        first_line = LineString(np.column_stack((frequencyList, CMfactorList)))
        second_line = LineString(np.column_stack((frequencyList, zeroList)))
        back_line = LineString(np.column_stack((frequencyList, backList)))

        intersection = first_line.intersection(second_line)
        backintersection = first_line.intersection(back_line)

        try:
            print(float(re.split(r"[-;,()\s]\s*", str(intersection))[2]))
            if intersection.geom_type == 'Point':
                if float(re.split(r"[-;,()\s]\s*", str(intersection))[2]) < float(re.split(r"[-;,()\s]\s*", str(backintersection))[2]):
                    firstCO = float(re.split(r"[-;,()\s]\s*", str(intersection))[2])
                    secondCO = 'N/A'
                elif float(re.split(r"[-;,()\s]\s*", str(intersection))[2]) > float(re.split(r"[-;,()\s]\s*", str(backintersection))[2]):
                    firstCO = 'N/A'
                    secondCO = float(re.split(r"[-;,()\s]\s*", str(intersection))[2])

            elif intersection.geom_type == 'MultiPoint':
                first_intersection = float(re.split(r"[-;,()\s]\s*", str(intersection))[2])
                second_intersection = float(re.split(r"[-;,()\s]\s*", str(intersection))[4])
                if first_intersection < second_intersection:
                    firstCO = first_intersection
                    secondCO = second_intersection
                elif first_intersection > second_intersection:
                    firstCO = second_intersection
                    secondCO = first_intersection

            else:
                firstCO = 'N/A'
                secondCO = 'N/A'

        except:
            firstCO = 'N/A'
            secondCO = 'N/A'
            print("[WARNING] [FITTING] [CROSSOVER] Crossing-over couldn't be computed")

        return firstCO, secondCO, intersection


    def cross_over(self,frequencyList, CMfactorList, firstcrossoverentry, secondcrossoverentry, markcrossover, MplCMWidget):
        firstCO, secondCO, intersection = self.cross_over_first(frequencyList, CMfactorList)
        try:
            if firstCO != 'N/A':
                if firstCO > 1000000:
                    firstcrossoverentry.setText(str(round((firstCO / 1000000), 2)) + " MHz")
                elif firstCO > 1000:
                    firstcrossoverentry.setText(str(int(firstCO / 1000)) + " kHz")
                else:
                    firstcrossoverentry.setText(str(int(firstCO)) + " Hz")
            else:
                firstcrossoverentry.setText('N/A')

            if secondCO != 'N/A':
                if secondCO > 1000000:
                    secondcrossoverentry.setText(str(round((secondCO / 1000000), 2)) + " MHz")
                elif secondCO > 1000:
                    secondcrossoverentry.setText(str(int(secondCO / 1000)) + " kHz")
                else:
                    secondcrossoverentry.setText(str(int(secondCO)) + " Hz")
            else:
                secondcrossoverentry.setText('N/A')

            if markcrossover.isChecked():
                if intersection.geom_type == 'MultiPoint':
                    MplCMWidget.canvas.axes.plot([firstCO, secondCO],[0,0], 'x')
                elif intersection.geom_type == 'Point':
                    MplCMWidget.canvas.axes.plot(*intersection.xy, 'x')
        except:
            print("[WARNING] [FITTING] [CROSSOVER] Crossing-over couldn't be saved")

## Data manipulation and excel importing / exporting ##
class Data():

    def __init__(self):
        # Instance Variable #
        self.frequency_list = []
        self.CMfactor_list = []
        self.CMfactor_errors_list = []
        self.DEPforce_list = []
        self.DEPforce_errors_list = []
        self.dataStatus_list = []
        self.dataValidation_list = []

        self.fittedFrequency_list = []
        self.fittedCMfactor_list = []
        self.fittedDEPforce_list = []

    def removeSelectedPoints(self, input_list):
        local_data_list = []
        for i in input_list:
            local_data_list.append(i)
        try:
            if len(self.dataStatus_list) > 2:
                for i in local_data_list:
                    j = local_data_list.index(i)
                    if self.dataStatus_list[j] == 'Disabled':
                        local_data_list[j] = 'null'

                local_data_list2 = [s for s in  local_data_list if s != 'null']
                return local_data_list2

            else:
                return input_list
                print('output list')
        except:
            print("================= [ERROR] =========================")
            print("[datamanagement] [Data] removeSelectedPoints method could not run")

# Loads data_01 from an excel file #
    def loadExcelData(self, file):
        self.frequency_list = []
        self.CMfactor_list = []
        self.CMfactor_errors_list = []
        self.DEPforce_list = []
        self.DEPforce_errors_list = []
        self.dataValidation_list = []
        self.dataStatus_list = []


        self.fittedFrequency_list = []
        self.fittedCMfactor_list = []
        self.fittedDEPforce_list = []

        wb = load_workbook(filename=file)
        ws = wb.active
        max_row = ws.max_row
        listWithLists = [[self.frequency_list, self.CMfactor_list, self.CMfactor_errors_list, self.DEPforce_list, self.DEPforce_errors_list, self.dataValidation_list, self.dataStatus_list, self.fittedFrequency_list, self.fittedCMfactor_list,self.fittedDEPforce_list], [1,2,3,4,5,6,7,8,9,10]]


        for i in listWithLists[1]:
            for j in range(2, max_row):
                if isinstance(ws.cell(row=j, column=i).value, float) == True:
                    listWithLists[0][i-1].append(ws.cell(row=j, column=i).value)
                elif isinstance(ws.cell(row=j, column=i).value, int) == True:
                    listWithLists[0][i-1].append(float(ws.cell(row=j, column=i).value))
                elif ws.cell(row=j, column=i).value != None:
                    listWithLists[0][i - 1].append(ws.cell(row=j, column=i).value)


        if len(self.dataValidation_list) < 2:
            for i in self.frequency_list:
                self.dataValidation_list.append("Normal")

        if len(self.dataStatus_list) < 2:
            for i in self.frequency_list:
                self.dataStatus_list.append("Enabled")

        #print(self.frequency_list, self.CMfactor_list, self.CMfactor_errors_list, self.DEPforce_list, self.DEPforce_errors_list, self.dataValidation_list, self.dataStatus_list)


    def saveExcelData(self, file, model, MainUI, data, fit):

        wb = Workbook()
        ws = wb.active

    ## Excel top-table setup ##
        ws['A1'] = 'Frequency (Hz)'
        ws['B1'] = 'Experimental CM factor'
        ws['C1'] = 'Experimental CM Factor errors'
        ws['D1'] = 'Experimental DEP force'
        ws['E1'] = 'Experimental DEP force errors'
        ws['F1'] = 'Status'
        ws['G1'] = 'Type'
        ws['H1'] = 'Fitted Frequency (Hz)'
        ws['I1'] = 'Fitted CMfactor'
        ws['J1'] = 'Fitted DEPforce'
        ws['K1'] = 'Name               '
        ws['L1'] = 'Value / Type'
        ws['M1'] = 'Error (+/-)'
        ws['N1'] = 'Unit     '

        self.autoStretchColumns(ws)
        self.setBackgroundColor(ws, top=1, bottom=1000, left=1, right=100, color="00EBF1DE")
        self.setBackgroundColor(ws, top=1, bottom=1, left=1, right=14, color='00C4D79B')
        self.setBorder(ws, top=1, bottom=1, left=1, right=14, color='000000', border_style='thick')

        firstCO, secondCO, garbage = fit.cross_over_first(self.fittedFrequency_list, self.fittedCMfactor_list)
        try:
            rsquare = fit.returnFitRsquare(fit.result)
        except:
            rsquare = 'None'
            print('No RSquare')

        if model == 'hopa':
            dict = {'Model':[model, ' '],
                    'Rsqare':[rsquare, ' '],
                    '1st CO frequency':[firstCO, 'Hz'],
                    '2nd CO frequency':[secondCO, 'Hz'],
                    'Buffer Conductivity': [fit.returnFittedParamtersVar2(fit.result, 'fitting_gen_buffer_cond'), 'S/m'],
                    'Buffer Permitivity': [fit.returnFittedParamtersVar2(fit.result, 'fitting_gen_buffer_perm'), 'ε'],
                    'Particle Conductivity': [fit.returnFittedParamtersVar2(fit.result, 'fitting_hopa_particle_cond'), 'S/m'],
                    'Particle Permitivity': [fit.returnFittedParamtersVar2(fit.result, 'fitting_hopa_particle_perm'), 'ε'],
                    'Particle radius': [fit.returnFittedParamtersVar2(fit.result, 'fitting_hopa_particle_radius'), 'μm']
                    }

        elif model == 'sish':
            dict = {'Model': [model, ' '],
                    'Rsqare': [rsquare, ' '],
                    '1st CO frequency':[firstCO, 'Hz'],
                    '2nd CO frequency':[secondCO, 'Hz'],
                    'Buffer Conductivity': [fit.returnFittedParamtersVar2(fit.result, 'fitting_gen_buffer_cond'), 'S/m'],
                    'Buffer Permitivity': [fit.returnFittedParamtersVar2(fit.result, 'fitting_gen_buffer_perm'), 'ε'],
                    'Particle Radius': [fit.returnFittedParamtersVar2(fit.result, 'fitting_sish_particle_radius'), 'μm'],
                    'Membrane Thickness': [fit.returnFittedParamtersVar2(fit.result, 'fitting_sish_membrane_thickness'), 'nm'],
                    'Membrane Conductivity': [fit.returnFittedParamtersVar2(fit.result, 'fitting_sish_membrane_cond'), 'S/m'],
                    'Membrane Permitivity': [fit.returnFittedParamtersVar2(fit.result, 'fitting_sish_membrane_perm'), 'ε'],
                    'Membrane Conductance': [MainUI.fitting_sish_membrane_conductance_value_entry.text(), 'S/m^2'],
                    'Membrane Capacitance': [MainUI.fitting_sish_membrane_capacitance_value_entry.text(), 'mF/m^2'],
                    'Cytoplasm Conductivity': [fit.returnFittedParamtersVar2(fit.result, 'fitting_sish_cytoplasm_cond'), 'S/m'],
                    'Cytoplasm Permitivity': [fit.returnFittedParamtersVar2(fit.result, 'fitting_sish_cytoplasm_perm'), 'ε']
                    }

        x = 2
        for i in dict.keys():
            ws.cell(row=x, column=11, value=i)
            if i == 'Membrane Conductance' or i == 'Membrane Capacitance' or i == 'Model' or i == 'Rsqare' or i == '1st CO frequency' or i == '2nd CO frequency':
                try:
                    ws.cell(row=x, column=12, value=float(dict.get(i)[0]))
                    ws.cell(row=x, column=13, value='None')
                except:
                    ws.cell(row=x, column=12, value=dict.get(i)[0])
                    ws.cell(row=x, column=13, value='None')
            else:
                try:
                    ws.cell(row=x, column=12, value=float(dict.get(i)[0][0]))
                except:
                    ws.cell(row=x, column=12, value=dict.get(i)[0][0])
                try:
                    ws.cell(row=x, column=13, value=float(dict.get(i)[0][1]))
                except:
                    ws.cell(row=x, column=13, value=dict.get(i)[0][1])

            ws.cell(row=x, column=14, value=dict.get(i)[1])
            x = x + 1
        self.setBorder(ws, top=2, bottom=x-1, left=11, right=14, color='000000', border_style='double')
        self.setTabelBackground(ws, top=2, bottom=x-1, left=11, right=14, color1='00C0C0C0', color2='00FFFFFF')


        list = [self.frequency_list, self.CMfactor_list, self.CMfactor_errors_list, self.DEPforce_list, self.DEPforce_errors_list, self.dataValidation_list, self.dataStatus_list]
        x = 1
        for i in list:
            y = 2
            for j in i:
                ws.cell(row=y, column=x, value=j)
                y = y + 1
            x = x + 1

        list2 = []
        for i in list:
            list2.append(len(i))
        self.setBorder(ws, top=2, bottom=max(list2)+1, left=1, right=7, color='000000', border_style='double')
        self.setTabelBackground(ws, top=2, bottom=max(list2)+1, left=1, right=7, color1='00C0C0C0', color2='00FFFFFF')

        list = [self.fittedFrequency_list, self.fittedCMfactor_list, self.fittedCMfactor_list]
        x = 8
        for i in list:
            y = 2
            for j in i:
                ws.cell(row=y, column=x, value=j)
                y = y + 1
            x = x + 1

        list2 = []
        for i in list:
            list2.append(len(i))
        self.setBorder(ws, top=2, bottom=max(list2)+1, left=8, right=x - 1, color='000000', border_style='double')
        self.setTabelBackground(ws, top=2, bottom=max(list2)+1, left=8, right=x-1, color1='00C0C0C0', color2='00FFFFFF')

    # Adding the charts#
        self.chartPlotExcel(ws, what_to_plot="CMfactor", exp_freq=data.frequency_list,fit_freq=data.fittedFrequency_list, name='CM factor scatter', yname='CM factor', position="O1")
        self.chartPlotExcel(ws, what_to_plot="DEPforce", exp_freq=data.frequency_list,fit_freq=data.fittedFrequency_list, name='DEP force scatter', yname='DEP force', position="O20")

        wb.save(filename=file)


    # Setting borders to a group of cells #
    def setBorder(self, ws, top, bottom, left, right, color, border_style):

        for i in range(top, bottom):
            border1 = Border(left=Side(border_style=border_style, color=color))
            ws.cell(row=i, column=left).border = border1
            border2 = Border(right=Side(border_style=border_style, color=color))
            ws.cell(row=i, column=right).border = border2

        for i in range(left, right):
            border1 = Border(top=Side(border_style=border_style, color=color))
            ws.cell(row=top, column=i).border = border1
            border2 = Border(bottom=Side(border_style=border_style, color=color))
            ws.cell(row=bottom, column=i).border = border2

        border = Border(left=Side(border_style=border_style, color=color), top=Side(border_style=border_style, color=color))
        ws.cell(row=top, column=left).border = border
        border = Border(left=Side(border_style=border_style, color=color), bottom=Side(border_style=border_style, color=color))
        ws.cell(row=bottom, column=left).border = border
        border = Border(top=Side(border_style=border_style, color=color), right=Side(border_style=border_style, color=color))
        ws.cell(row=top, column=right).border = border
        border = Border(bottom=Side(border_style=border_style, color=color), right=Side(border_style=border_style, color=color))
        ws.cell(row=bottom, column=right).border = border

    def setBackgroundColor(self, ws, top, bottom, left, right, color):
        fill = PatternFill("solid", fgColor=color)
        for i in range(top, bottom+1):
            for j in range(left, right+1):
                ws.cell(row=i, column=j).fill = fill

    def setTabelBackground(self, ws, top, bottom, left, right, color1, color2):
        for i in range(top, bottom+1):
            if (i % 2) == 0:
                self.setBackgroundColor(ws, top=i, bottom=i, left=left, right=right, color=color1)
            else:
                self.setBackgroundColor(ws, top=i, bottom=i, left=left, right=right, color=color2)

    def autoStretchColumns(self, ws):
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column name
            # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based)
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1
            ws.column_dimensions[column].width = adjusted_width

    def chartPlotExcel(self, ws ,what_to_plot, exp_freq, fit_freq, name, yname, position):

        chart = ScatterChart()
        chart.title = name
        #chart.style = 13
        chart.x_axis.title = 'Frequency (Hz)'
        chart.y_axis.title = yname
        chart.height = 10
        chart.width = 20

        chart.x_axis.scaling.logBase = 10
        chart.x_axis.scaling.min = min(fit_freq)
        chart.x_axis.scaling.max = max(fit_freq)

        if what_to_plot == "CMfactor":
            i=2
            j=9
        elif what_to_plot == "DEPforce":
            i=3
            j=10

        xvalues1 = Reference(ws, min_col=1 , min_row=2, max_row=2 + len(exp_freq))
        values1 = Reference(ws, min_col=i, min_row=2, max_row=2 + len(exp_freq))
        series1 = Series(values=values1, xvalues=xvalues1, title='Experimental data_01')

        xvalues2 = Reference(ws, min_col=8, min_row=2, max_row=2 + len(fit_freq))
        values2 = Reference(ws, min_col=j, min_row=2, max_row=2 + len(fit_freq))
        series2 = Series(values=values2, xvalues=xvalues2, title='Fit')

        chart.series.append(series1)
        chart.series.append(series2)

        ws.add_chart(chart, position)